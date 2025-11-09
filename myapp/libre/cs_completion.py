# myapp/libre/cs_completion.py
import os, subprocess, tempfile
from decimal import Decimal
from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook

TEMPLATE_PATH = os.path.join(
    settings.BASE_DIR, "myapp", "cert_templates", "Completion_Certificate.xlsx"
)

def _set_input(ws, key, value):
    k = (key or "").strip().lower()
    for row in ws.iter_rows(min_row=1, max_col=2):
        a = (row[0].value or "").strip().lower()
        if a == k:
            row[1].value = "" if value is None else str(value)
            return True
    return False

def _fmt_student_name(case):
    parts = []
    if getattr(case, "first_name", None):
        parts.append(case.first_name.strip())
    mi = (getattr(case, "middle_initial", "") or "").strip().rstrip(".")
    if mi:
        parts.append(f"{mi}.")
    if getattr(case, "last_name", None):
        parts.append(case.last_name.strip())
    ext = (getattr(case, "extension_name", "") or "").strip()
    if ext:
        parts.append(ext)
    return " ".join(parts)

def _fmt_date_short(dt):
    # 11/02/2025 style (MM/DD/YYYY), no leading zero for month
    local = timezone.localtime(dt)
    s = local.strftime("%m/%d/%Y")
    return s.lstrip("0")  # "11/02/2025" -> "11/02/2025", "02/05/2025" -> "2/05/2025"

def _fmt_date_long(d):
    # November 10, 2025
    local = timezone.localtime(d) if hasattr(d, "tzinfo") else d
    return local.strftime("%B %-d, %Y") if os.name != "nt" else local.strftime("%B %d, %Y").replace(" 0", " ")

def _fmt_hours(x):
    # Show 20 not 20.0
    if isinstance(x, Decimal):
        x = x.normalize()
    try:
        f = float(x)
        if f.is_integer():
            return str(int(f))
        return str(f)
    except Exception:
        return str(x)

def build_cs_completion_pdf(case_obj, osa_head_name):
    """
    Fill `inputs` then export 'Completion Certificate' to PDF via LibreOffice.
    Keys required in A-column: student_name, student_id, program, date, start_date, end_date, osa_head, hours
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    # derive start/end from logs (fallbacks if none)
    logs_qs = case_obj.logs.order_by("check_in_at")
    first_in  = logs_qs.first().check_in_at if logs_qs.exists() else timezone.now()
    last_out_qs = case_obj.logs.exclude(check_out_at__isnull=True).order_by("-check_out_at")
    last_dt = (last_out_qs.first().check_out_at if last_out_qs.exists()
               else logs_qs.last().check_in_at if logs_qs.exists()
               else timezone.now())

    payload = {
        "student_name": _fmt_student_name(case_obj),
        "student_id":   case_obj.student_id,
        "program":      getattr(case_obj, "program_course", ""),
        "date":         _fmt_date_short(timezone.now()),
        "start_date":   _fmt_date_long(first_in),
        "end_date":     _fmt_date_long(last_dt),
        "osa_head":     osa_head_name or "",
        # For a completion cert, "required __ hours" typically = total_required_hours
        # (which should equal hours_completed at completion time)
        "hours":        _fmt_hours(getattr(case_obj, "total_required_hours", getattr(case_obj, "hours_completed", ""))),
    }

    tmpdir = tempfile.mkdtemp(prefix="cs_complete_")
    xlsx_out = os.path.join(tmpdir, "completion_work.xlsx")

    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws_inputs = wb["inputs"]

    for k, v in payload.items():
        if not _set_input(ws_inputs, k, v):
            raise KeyError(f"Key '{k}' not found in inputs sheet of {TEMPLATE_PATH}")

    # Hide inputs & activate target
    ws_inputs.sheet_state = "hidden"
    target_name = "Completion Certificate"
    if target_name not in wb.sheetnames:
        raise KeyError(f"Sheet '{target_name}' not found in template.")
    wb.active = wb.sheetnames.index(target_name)

    # Avoid stale Print_Area warnings; auto-fit to used cells
    try:
        ws_target = wb[target_name]
        ws_target.print_area = ws_target.calculate_dimension()
        wb.defined_names.delete("Print_Area")
    except Exception:
        pass

    wb.save(xlsx_out)

    lo = getattr(settings, "LIBREOFFICE_BIN", "soffice")
    cmd = [lo, "--headless", "--nologo", "--nofirststartwizard",
           "--convert-to", "pdf", "--outdir", tmpdir, xlsx_out]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if proc.returncode != 0:
        err = proc.stderr.decode(errors="ignore") or proc.stdout.decode(errors="ignore")
        raise RuntimeError(f"LibreOffice export failed: {err}")

    pdf_path = os.path.join(tmpdir, "completion_work.pdf")
    nice = os.path.join(tmpdir, f"Completion-{case_obj.student_id}.pdf")
    if os.path.exists(pdf_path):
        os.replace(pdf_path, nice)
        pdf_path = nice
    return pdf_path
