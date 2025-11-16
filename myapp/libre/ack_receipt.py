
import os
import tempfile
import subprocess

from django.conf import settings
from django.utils import timezone
from openpyxl import load_workbook

TEMPLATE_PATH = os.path.join(
    settings.BASE_DIR,
    "myapp",
    "cert_templates",
    "Acknowledgement-Receipt-Form.xlsx",
)


def _fmt_name(req):
    parts = [req.first_name, req.middle_name, req.surname, req.extension]
    name = " ".join(p for p in parts if p).strip()
    # collapse duplicate spaces and uppercase
    return " ".join(name.split()).upper()


def _fmt_timestamp(dt):
    # Example: 10/08/2025 8:57:32PM (no space before AM/PM)
    local = timezone.localtime(dt)
    s = local.strftime("%m/%d/%Y %I:%M:%S%p")
    return s.lstrip("0")


def _set_named_value(wb, key, value):
    """
    Write value into a Named Range (like 'student_name', 'program', etc.).
    Your screenshot shows names: student_name, student_id, program, proof, reason,
    year_level, years_of_stay, timestamp, osa_head.
    """
    key = (key or "").strip().lower()
    if not key:
        return False

    # openpyxl: wb.defined_names.definedName is a list of DefinedName objects
    for defined in wb.defined_names.definedName:
        if defined.name.lower() == key:
            # Named range may technically refer to multiple cells,
            # but in your case each name is a single cell.
            for sheet_name, coord in defined.destinations:
                ws = wb[sheet_name]
                ws[coord].value = value
            return True
    return False


def build_ack_pdf_bytes(request_obj, admin_name_upper):
    """
    Fill Named Ranges in the Acknowledgement Receipt template and export to PDF.

    Returns: PDF bytes (not a file path).
    """
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    # temp dir for this one export
    tmpdir = tempfile.mkdtemp(prefix="ack_")
    xlsx_out = os.path.join(tmpdir, "ack_work.xlsx")

    # ---- 1. Load workbook and fill named ranges ----
    wb = load_workbook(TEMPLATE_PATH, data_only=False)

    data = {
        "student_name": _fmt_name(request_obj),
        "student_id": request_obj.student_number,
        "year_level": request_obj.year_level,
        "program": request_obj.program,
        "years_of_stay": request_obj.inclusive_years,
        "timestamp": _fmt_timestamp(timezone.now()),
        "proof": (
            request_obj.get_document_type_display()
            if hasattr(request_obj, "get_document_type_display")
            else request_obj.document_type
        ),
        "reason": (
            request_obj.get_reason_display()
            if hasattr(request_obj, "get_reason_display")
            else request_obj.reason
        ),
        "osa_head": admin_name_upper,
    }

    for k, v in data.items():
        _set_named_value(wb, k, v)

    # Optional but still fine: hide inputs / set visible sheet
    if "inputs" in wb.sheetnames:
        ws_inputs = wb["inputs"]
        ws_inputs.sheet_state = "hidden"
    if "Acknowledgement Receipt" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Acknowledgement Receipt")

    wb.save(xlsx_out)

    # ---- 2. Export to PDF via LibreOffice headless ----
    outdir = tmpdir
    lo = getattr(settings, "LIBREOFFICE_BIN", "soffice")
    cmd = [
        lo,
        "--headless",
        "--nologo",
        "--nofirststartwizard",
        "--convert-to",
        "pdf",
        "--outdir",
        outdir,
        xlsx_out,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if proc.returncode != 0:
        raise RuntimeError(
            f"LibreOffice export failed: "
            f"{proc.stderr.decode(errors='ignore') or proc.stdout.decode(errors='ignore')}"
        )

    pdf_path = os.path.join(outdir, "ack_work.pdf")
    if not os.path.exists(pdf_path):
        raise RuntimeError("Expected PDF not found after LibreOffice export.")

    # Read bytes and clean up
    with open(pdf_path, "rb") as f:
        pdf_bytes = f.read()

    # (You can optionally delete the temp dir; OS usually cleans /tmp eventually)
    # import shutil; shutil.rmtree(tmpdir, ignore_errors=True)

    return pdf_bytes