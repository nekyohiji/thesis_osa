# myapp/libre/cs_agreement.py
import os, subprocess, tempfile
from django.conf import settings
from openpyxl import load_workbook
from myapp.utils import _get_osa_head_name, _get_osa_head_title

TEMPLATE_PATH = os.path.join(
    settings.BASE_DIR, "myapp", "cert_templates", "Agreement_Certificate.xlsx"
)

def _set_input(ws, key, value):
    """Find `key` in column A (case-insensitive exact match) and write `value` to column B."""
    k = (key or "").strip().lower()
    for row in ws.iter_rows(min_row=1, max_col=2):
        a = (row[0].value or "").strip().lower()
        if a == k:
            row[1].value = "" if value is None else str(value)
            return True
    return False

def _fmt_student_name(case):
    # Build "First M. Last Ext" with normalized spacing; preserve casing.
    parts = []
    if getattr(case, "first_name", None):
        parts.append(case.first_name.strip())
    mi = (getattr(case, "middle_initial", "") or "").strip().rstrip(".")
    if mi:
        parts.append(f"{mi}.")
    if getattr(case, "last_name", None):
        parts.append(case.last_name.strip())
    ext = (getattr(case, "extension_name", "") or "").strip()
    if ext:
        parts.append(ext)
    name = " ".join(p for p in parts if p)
    return " ".join(name.split())

def build_cs_agreement_pdf(case_obj, osa_head_name=None, osa_head_title=None):
    if not os.path.exists(TEMPLATE_PATH):
        raise FileNotFoundError(f"Template not found: {TEMPLATE_PATH}")

    if osa_head_name is None:
        osa_head_name = _get_osa_head_name()
    if osa_head_title is None:
        osa_head_title = _get_osa_head_title()

    tmpdir = tempfile.mkdtemp(prefix="cs_agree_")
    xlsx_out = os.path.join(tmpdir, "agreement_work.xlsx")

    wb = load_workbook(TEMPLATE_PATH, data_only=False)
    ws_inputs = wb["inputs"]

    payload = {
        "student_name": _fmt_student_name(case_obj),
        "program":      getattr(case_obj, "program_course", ""),
        "osa_head":     osa_head_name or "",
        "title":        osa_head_title or "",   
    }

    for k, v in payload.items():
        ok = _set_input(ws_inputs, k, v)
        if not ok:
            raise KeyError(f"Key '{k}' not found in inputs sheet A-column of {TEMPLATE_PATH}")

    ws_inputs.sheet_state = "hidden"
    if "Agreement Certificate" in wb.sheetnames:
        wb.active = wb.sheetnames.index("Agreement Certificate")

    wb.save(xlsx_out)

    lo = getattr(settings, "LIBREOFFICE_BIN", "soffice")
    cmd = [
        lo, "--headless", "--nologo", "--nofirststartwizard",
        "--convert-to", "pdf", "--outdir", tmpdir, xlsx_out,
    ]
    proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
    if proc.returncode != 0:
        err = proc.stderr.decode(errors="ignore") or proc.stdout.decode(errors="ignore")
        raise RuntimeError(f"LibreOffice export failed: {err}")

    pdf_path = os.path.join(tmpdir, "agreement_work.pdf")
    nice_name = f"Agreement-{case_obj.student_id}.pdf"
    nice_path = os.path.join(tmpdir, nice_name)
    if os.path.exists(pdf_path):
        os.replace(pdf_path, nice_path)
        pdf_path = nice_path

    return pdf_path
