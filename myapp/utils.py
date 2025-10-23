from django.core.mail import send_mail
from django.conf import settings
import os, json, datetime, tempfile, subprocess, uuid, sys, importlib
from django.conf import settings
from django.core.files import File
from django.db import connection
from openpyxl import load_workbook
from .models import GoodMoralRequest, UserAccount
from django.core.mail import EmailMultiAlternatives
from django.utils.html import escape
import mimetypes
from pathlib import Path
import os, json, subprocess, tempfile, shutil
from pathlib import Path
from django.conf import settings
from openpyxl import load_workbook
from .models import UserAccount
from django.utils.timezone import localtime
from django.utils.html import strip_tags, escape
from textwrap import dedent
from django.http import JsonResponse
import json
import secrets
from django.core.mail import EmailMultiAlternatives, BadHeaderError
from django.utils import timezone
import unicodedata
import re
from threading import Thread
from django.core.mail import send_mail as _send_mail
from decimal import Decimal

EMAIL_MAX_ATTACHMENT_SIZE = getattr(settings, "EMAIL_MAX_ATTACHMENT_SIZE", 10_000_000)  # ~5 MB

def _ordinal(n: int) -> str:
    if n == 1: return "first"
    if n == 2: return "second"
    if n == 3: return "third"
    return f"{n}th"

def send_violation_email(request, violation, student, violation_count=None, settlement_type=None, declined=False):
    """
    Sends an email to the student.
    - Approved: attaches evidence images (no links in body).
    - Declined: simple notice.
    """
    to_addr = [student.email]
    from_addr = settings.DEFAULT_FROM_EMAIL

    if declined:
        subject = "TUPC OSA: Violation Report Declined"
        text_body = (
            f"Dear {student.first_name} {student.last_name},\n\n"
            f"The violation report filed under your name dated {violation.violation_date} "
            f"for '{violation.get_violation_type_display()}' has been reviewed and lifted by the Office of Student Affairs.\n\n"
            f"No further action is required on your part.\n\n"
            f"Thank you."
        )
        html_body = f"""
            <p>Dear {escape(student.first_name)} {escape(student.last_name)},</p>
            <p>The violation report filed under your name dated {escape(str(violation.violation_date))} for
            "<strong>{escape(violation.get_violation_type_display())}</strong>" has been reviewed and <strong>declined</strong>
            by the Office of Student Affairs.</p>
            <p>No further action is required on your part.</p>
            <p>Thank you.</p>
        """
        msg = EmailMultiAlternatives(subject, text_body, from_addr, to_addr)
        msg.attach_alternative(html_body, "text/html")
        msg.send(fail_silently=False)
        return

    # --- Approved path (attachments only, no links) ---
    ord_str = _ordinal(int(violation_count or 1))
    subject = f"TUPC OSA: Notice of {ord_str.capitalize()} Violation"

    evidence = []
    if getattr(violation, "evidence_1", None):
        evidence.append(("Evidence 1", violation.evidence_1))
    if getattr(violation, "evidence_2", None):
        evidence.append(("Evidence 2", violation.evidence_2))

    # Body text without any URLs
    if evidence:
        evidence_line_text = "Evidence files are attached to this email."
        evidence_line_html = "<p>Evidence files are attached to this email.</p>"
    else:
        evidence_line_text = "No evidence files were attached."
        evidence_line_html = "<p>No evidence files were attached.</p>"

    text_body = (
        f"Dear {student.first_name} {student.last_name},\n\n"
        f"You have committed your {ord_str} violation on {violation.violation_date}.\n"
        f"Violation Type: {violation.get_violation_type_display()}\n"
        f"You are required to settle the {settlement_type} at the Office of Student Affairs IMMEDIATELY.\n\n"
        f"{evidence_line_text}\n\n"
        f"Please visit the Office of Student Affairs for more details.\n\n"
        f"Thank you."
    )

    html_body = (
        f"<p>Dear {escape(student.first_name)} {escape(student.last_name)},</p>"
        f"<p>You have committed your <strong>{escape(ord_str)}</strong> violation on {escape(str(violation.violation_date))}.<br>"
        f"Violation Type: <strong>{escape(violation.get_violation_type_display())}</strong><br>"
        f"You are required to settle the <strong>{escape(settlement_type)}</strong> at the Office of Student Affairs IMMEDIATELY.</p>"
        f"{evidence_line_html}"
        f"<p>Please visit the Office of Student Affairs for more details.</p>"
        f"<p>Thank you.</p>"
    )

    msg = EmailMultiAlternatives(subject, text_body, from_addr, to_addr)
    msg.attach_alternative(html_body, "text/html")

    # Attach images (respecting size cap)
    for label, f in evidence:
        try:
            f.open("rb")
            size = getattr(f.file, "size", None) or getattr(f, "size", None)
            if size is None or size <= EMAIL_MAX_ATTACHMENT_SIZE:
                content = f.read()
                guessed = mimetypes.guess_type(f.name)[0] or "application/octet-stream"
                filename = f.name.split("/")[-1]
                msg.attach(filename, content, guessed)
            # else: silently skip oversized files (no links in body)
        except Exception:
            # Skip if storage/backing fails; body already says "attached" or "no evidence"
            pass
        finally:
            try:
                f.close()
            except Exception:
                pass

    msg.send(fail_silently=False)

def _norm(s: str) -> str:
    """Lowercase, strip accents, keep letters/digits only."""
    if not s:
        return ""
    s = s.strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", s)

def build_student_email(first_name: str, last_name: str, extension_name: str = "", domain: str = None) -> str:
    """firstname.lastname{ext}@gsfe.tupcavite.edu.ph (domain overridable via settings)."""
    domain = domain or getattr(settings, "STUDENT_EMAIL_DOMAIN", "gsfe.tupcavite.edu.ph")
    first = _norm(first_name)
    last  = _norm(last_name)
    ext   = _norm(extension_name)
    if not first or not last:
        return ""
    local = f"{first}.{last}{ext}"
    return f"{local}@{domain}"

def send_violation_notice(violation, to_email: str, *, max_history: int = 10) -> tuple[bool, str]:
    """
    Email the student about the recorded violation, including:
      - total number of violations for this student_id
      - recent history (what & when), up to `max_history`
    Returns (ok, message_or_error).
    """
    if not to_email:
        return False, "Missing student email."
    if not getattr(settings, "SEND_VIOLATION_EMAILS", True):
        return False, "Email sending disabled by settings."

    # --- Build history for context (keyed by student_id) ---
    from .models import Violation  # local import to avoid circulars
    history_qs = (Violation.objects
                  .filter(student_id=violation.student_id)
                  .order_by("-violation_date", "-violation_time", "-created_at"))
    total_count = history_qs.count()
    recent = list(history_qs[:max_history])

    def _fmt_dt(v):
        d = v.violation_date.strftime("%Y-%m-%d")
        t = v.violation_time.strftime("%H:%M")
        return f"{d} {t}"

    history_lines_text = "\n".join(
        f"• {_fmt_dt(v)} — {v.violation_type} [{v.severity}] (Status: {v.status})"
        for v in recent
    ) or "• No prior records found."

    history_rows_html = "".join(
        f"<tr><td>{_fmt_dt(v)}</td><td>{v.violation_type}</td><td>{v.severity}</td><td>{v.status}</td></tr>"
        for v in recent
    ) or "<tr><td colspan='4'>No prior records found.</td></tr>"

    # --- Current record summary ---
    subject = "TUPC OSA: Violation Recorded"
    recorded_at_local = timezone.localtime(violation.created_at).strftime("%Y-%m-%d %H:%M")
    vdate = violation.violation_date.strftime("%Y-%m-%d")
    vtime = violation.violation_time.strftime("%H:%M")

    text_body = (
        f"Good day, {violation.first_name} {violation.last_name},\n\n"
        f"A violation record has been created in the OSA system.\n\n"
        f"Student ID: {violation.student_id}\n"
        f"Program: {violation.program_course}\n"
        f"Violation: {violation.violation_type} (Severity: {violation.severity})\n"
        f"Date/Time of Violation: {vdate} {vtime}\n"
        f"Recorded By: {violation.guard_name or 'OSA Admin'}\n"
        f"Status: {violation.status}\n"
        f"Recorded At: {recorded_at_local}\n\n"
        f"—\n"
        f"Total Violations on Record: {total_count}\n"
        f"Recent History (up to {max_history}):\n"
        f"{history_lines_text}\n\n"
        "If you believe this is incorrect, please contact the Office of Student Affairs.\n"
        "This is an automated message. Please do not reply to this email."
    )

    html_body = f"""
    <p>Good day, <strong>{violation.first_name} {violation.last_name}</strong>,</p>
    <p>A violation record has been created in the OSA system.</p>
    <table style="border-collapse:collapse">
      <tr><td><strong>Student ID</strong></td><td>{violation.student_id}</td></tr>
      <tr><td><strong>Program</strong></td><td>{violation.program_course}</td></tr>
      <tr><td><strong>Violation</strong></td><td>{violation.violation_type} (Severity: {violation.severity})</td></tr>
      <tr><td><strong>Date/Time of Violation</strong></td><td>{vdate} {vtime}</td></tr>
      <tr><td><strong>Recorded By</strong></td><td>{violation.guard_name or 'OSA Admin'}</td></tr>
      <tr><td><strong>Status</strong></td><td>{violation.status}</td></tr>
      <tr><td><strong>Recorded At</strong></td><td>{recorded_at_local}</td></tr>
    </table>

    <hr style="margin:16px 0;border:none;border-top:1px solid #ddd" />

    <p><strong>Total Violations on Record:</strong> {total_count}</p>
    <p><strong>Recent History (up to {max_history}):</strong></p>
    <table style="border-collapse:collapse;width:100%">
      <thead>
        <tr>
          <th style="text-align:left">Date/Time</th>
          <th style="text-align:left">Violation</th>
          <th style="text-align:left">Severity</th>
          <th style="text-align:left">Status</th>
        </tr>
      </thead>
      <tbody>
        {history_rows_html}
      </tbody>
    </table>

    <p style="color:#666">This is an automated message. Please do not reply to this email.</p>
    """

    from_email = getattr(settings, "DEFAULT_FROM_EMAIL", "no-reply@tupcavite.edu.ph")
    try:
        msg = EmailMultiAlternatives(subject, text_body, from_email, [to_email])
        msg.attach_alternative(html_body, "text/html")
        msg.send(fail_silently=False)
        return True, f"Email sent to {to_email}"
    except BadHeaderError:
        return False, "Invalid header found."
    except Exception as e:
        return False, f"Email failed: {e}"



#######################################

def send_status_email(to_email, approved=True, reason=None):
    subject = "Good Moral Certificate Request Status"
    if approved:
        message = "Your Good Moral Certificate request has been approved."
    else:
        message = f"Your request has been declined. Reason: {reason or 'Not specified'}"

    send_mail(
        subject,
        message,
        settings.DEFAULT_FROM_EMAIL,
        [to_email],
        fail_silently=False,
    )
    
# ---------- helpers ----------
def _clean_suffix(ext) -> str:
    if not ext:
        return ""
    raw = str(ext).strip()
    if not raw:
        return ""
    norm = "".join(ch for ch in raw if ch.isalnum()).upper()
    if norm in {"NA", "NONE", "NULL", "NAN"} or raw in {"-", "--"}:
        return ""
    return raw

def _format_student_name(req):
    if getattr(req, "student_name", None):
        return (req.student_name or "").strip()
    parts = [(getattr(req, "first_name", "") or "").strip()]
    mi = (getattr(req, "middle_name", "") or "").strip()
    if mi:
        parts.append(f"{mi[0].upper()}.")
    parts.append((getattr(req, "surname", "") or "").strip())
    ext_clean = _clean_suffix(getattr(req, "ext", ""))
    if ext_clean:
        parts.append(ext_clean)
    return " ".join(p for p in parts if p)

def _status_for_excel(raw: str) -> str:
    s = (raw or "").strip().lower()
    if s in {"current", "current student", "enrolled"}: return "Current Student"
    if s in {"former", "former student"}:               return "Former Student"
    if "grad" in s or "alum" in s:                      return "Graduate"
    return "Graduate"

def _fmt_yyyy_mm_dd(dt):
    return "" if not dt else dt.strftime("%Y-%m-%d")

def _get_osa_head_name():
    name = (UserAccount.objects
            .filter(role__iexact='admin', is_active=True)
            .order_by('-created_at', '-id')
            .values_list('full_name', flat=True)
            .first())
    return name.strip() if name else "BEVERLY M. DE VEGA"

def _find_soffice() -> str | None:
    for p in (
        os.environ.get("SOFFICE_BIN"),
        shutil.which("soffice"),
        shutil.which("libreoffice"),
        "/usr/bin/soffice", "/usr/bin/libreoffice", "/snap/bin/libreoffice",
    ):
        if p and os.path.exists(p):
            return p
    return None

# ---------- Excel filling (named ranges) ----------
def _fill_named_ranges(template_xlsx: Path, values: dict,
                       inputs_sheet="inputs", cert_sheet="GMF") -> Path:
    wb = load_workbook(str(template_xlsx))
    # write to single-cell defined names
    for key, val in values.items():
        dn = wb.defined_names.get(key)
        if not dn:
            continue
        for sheet_name, coord in dn.destinations:
            ws = wb[sheet_name]
            cell = coord.split(":")[0]  # top-left if it was a range
            ws[cell].value = val

    # force recalculation in LO on open
    try:
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    # show GMF, hide inputs, activate GMF
    if inputs_sheet in wb.sheetnames:
        wb[inputs_sheet].sheet_state = "hidden"
    if cert_sheet in wb.sheetnames:
        wb[cert_sheet].sheet_state = "visible"
        wb.active = wb.index(wb[cert_sheet])

    out_xlsx = Path(tempfile.gettempdir()) / f"gmf_fill_{next(tempfile._get_candidate_names())}.xlsx"
    wb.save(str(out_xlsx))
    return out_xlsx

# ---------- main entry ----------
def generate_gmf_pdf(req) -> bytes:
    template = Path(settings.GMF_TEMPLATE_PATH)
    if not template.exists():
        raise FileNotFoundError(f"GMF template not found at {template}")

    values = {
        "student_name":   _format_student_name(req),
        "program":        (getattr(req, "program", "") or "").strip(),
        "sex":            (getattr(req, "sex", "") or "").strip().upper(),  # Excel lowers it anyway
        "status":         _status_for_excel(getattr(req, "status", "")),
        # 👇 map to your model field names
        "years_of_stay":  (getattr(req, "inclusive_years", "") or "").strip(),
        "admission_date": (getattr(req, "date_admission", "") or "").strip(),
        "date_graduated": _fmt_yyyy_mm_dd(getattr(req, "date_graduated", None)),
        "purpose":        (getattr(req, "purpose", "") or "").strip(),
        "purpose_other":  (getattr(req, "other_purpose", "") or "").strip(),
        "osahead":        _get_osa_head_name(),
    }

    soffice = _find_soffice()
    if not soffice:
        raise FileNotFoundError(
            "LibreOffice 'soffice' not found. Install libreoffice-calc/writer "
            "or set SOFFICE_BIN to its path."
        )

    filled_xlsx = _fill_named_ranges(template, values, inputs_sheet="inputs", cert_sheet="GMF")

    out_dir = Path(tempfile.mkdtemp())
    try:
        cmd = [
            soffice, "--headless",
            "--convert-to", "pdf:calc_pdf_Export",
            "--outdir", str(out_dir),
            str(filled_xlsx),
        ]
        proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=120)
        pdf_path = out_dir / (filled_xlsx.stem + ".pdf")
        if proc.returncode != 0 or not pdf_path.exists():
            raise RuntimeError(
                "LibreOffice export failed.\n"
                f"CMD: {' '.join(cmd)}\n"
                f"STDOUT:\n{proc.stdout.decode(errors='ignore')}\n"
                f"STDERR:\n{proc.stderr.decode(errors='ignore')}"
            )
        return pdf_path.read_bytes()
    finally:
        try: os.remove(filled_xlsx)
        except Exception: pass
        shutil.rmtree(out_dir, ignore_errors=True)


#######################

def format_full_name(first: str, middle: str | None, last: str, extension: str | None = None, use_middle_initial: bool = False) -> str:
    first = (first or "").strip()
    mid   = (middle or "").strip()
    last  = (last or "").strip()
    ext   = (extension or "").strip()
    if mid:
        mid = f"{mid[0]}." if use_middle_initial else mid
        name = f"{first} {mid} {last}"
    else:
        name = f"{first} {last}"
    if ext:
        name = f"{name}, {ext}"
    return name

def _from_email() -> str:
    return getattr(settings, "DEFAULT_FROM_EMAIL", None) or getattr(settings, "EMAIL_HOST_USER", "")

def send_clearance_confirmation(obj) -> None:
    subject = "TUPC-Cavite OSA — Clearance Request Received"
    from_email = _from_email()

    # routing
    send_user_copy = getattr(settings, "OSA_SEND_USER_COPY", True)
    osa_inbox = getattr(settings, "OSA_INBOX", None)
    to = [obj.email] if (obj.email and send_user_copy) else []
    bcc = [osa_inbox] if osa_inbox else []
    reply_to = [osa_inbox] if osa_inbox else None

    submitted_local = localtime(obj.created_at)
    full_name = format_full_name(obj.first_name, obj.middle_name, obj.last_name, obj.extension)

    # ---------- Plain-text (nice labels + newlines) ----------
    text = dedent(f"""
        Clearance Request Received
        Thank you for your submission. Here are the details you provided:

        Name: {full_name}
        Email: {obj.email}
        Contact: {obj.contact}
        Student No.: {obj.student_number}
        Program: {obj.program}
        Year Level: {obj.year_level}
        Client Type: {obj.client_type}
        Stakeholder: {obj.stakeholder}
        Purpose: {obj.purpose}
        Submitted: {submitted_local:%B %d, %Y %I:%M %p}
    """).strip()

    # ---------- HTML (escaped vars, same content) ----------
    e = lambda s: escape(s or "")
    html = f"""
    <div style="font-family:system-ui,Segoe UI,Roboto,Arial,sans-serif;line-height:1.45">
      <h2 style="margin:0 0 12px">Clearance Request Received</h2>
      <p style="margin:0 0 12px">Thank you for your submission. Here are the details you provided:</p>
      <table cellpadding="6" cellspacing="0" style="border-collapse:collapse">
        <tr><td><b>Name</b></td><td>{e(full_name)}</td></tr>
        <tr><td><b>Email</b></td><td>{e(obj.email)}</td></tr>
        <tr><td><b>Contact</b></td><td>{e(obj.contact)}</td></tr>
        <tr><td><b>Student No.</b></td><td>{e(obj.student_number)}</td></tr>
        <tr><td><b>Program</b></td><td>{e(obj.program)}</td></tr>
        <tr><td><b>Year Level</b></td><td>{e(obj.year_level)}</td></tr>
        <tr><td><b>Client Type</b></td><td>{e(obj.client_type)}</td></tr>
        <tr><td><b>Stakeholder</b></td><td>{e(obj.stakeholder)}</td></tr>
        <tr><td><b>Purpose</b></td><td>{e(obj.purpose)}</td></tr>
        <tr><td><b>Submitted</b></td><td>{submitted_local:%B %d, %Y %I:%M %p}</td></tr>
      </table>
    </div>
    """.strip()

    msg = EmailMultiAlternatives(subject, text, from_email, to, bcc=bcc, reply_to=reply_to)
    msg.attach_alternative(html, "text/html")
    msg.send(fail_silently=False)

#############################
OTP_TTL_MINUTES = 10

def json_ok(message="ok", **extra):
    payload = {"status": "ok", "message": message}
    payload.update(extra)
    return JsonResponse(payload, status=200)

def json_err(message, *, code="ERROR", status=400, **extra):
    payload = {"status": "error", "code": code, "message": message}
    payload.update(extra)
    return JsonResponse(payload, status=status)

def safe_body(request):
    try:
        return json.loads(request.body or "{}")
    except json.JSONDecodeError:
        raise ValueError("Invalid JSON payload.")

def gen_otp():
    # 6-digit, cryptographically secure
    return "".join(secrets.choice("0123456789") for _ in range(6))


#################################
def send_mail_async(*args, **kwargs):
    def _job():
        try:
            _send_mail(*args, **kwargs)
        except Exception:
            pass
    Thread(target=_job, daemon=True).start()
    
    
def no_store(response):
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


##################################
CS_EXEMPT_TYPES = {"Attempt Fraternity"}  # keys must match Violation.violation_type

def compute_cs_topup_for_minor(violation_type: str, approved_count_of_same_type: int) -> Decimal:
    if violation_type in CS_EXEMPT_TYPES:
        return Decimal("0")
    if approved_count_of_same_type == 2:
        return Decimal("20")
    if approved_count_of_same_type == 3:
        return Decimal("10")
    return Decimal("0")  # 1st or 4th+ => no hours
